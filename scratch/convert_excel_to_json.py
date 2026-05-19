#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


OVERVIEW_SHEET = "总览表_录入表"
USAGE_SHEET = "使用说明"


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text or None


def as_int(value: Any) -> int | None:
    text = clean(value)
    if text is None:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def build_header_index(row: tuple[Any, ...]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, cell in enumerate(row):
        key = clean(cell)
        if key:
            result[key] = index
    return result


def cell(row: tuple[Any, ...], header_index: dict[str, int], name: str) -> Any:
    index = header_index.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


@dataclass
class WarningItem:
    level: str
    scope: str
    message: str
    row_number: int | None = None
    line_number: int | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "level": self.level,
            "scope": self.scope,
            "message": self.message,
            "row_number": self.row_number,
            "line_number": self.line_number,
        }


STANDALONE_INDEX_RE = re.compile(r"^[（(]?\d+[)）.]?$|^[①-⑳]$|^\d+[、.]$")
NUMBER_PREFIX_RE = re.compile(r"^[（(]?\d+[)）.、]\s*")
JAPANESE_SCRIPT_RE = re.compile(r"[ぁ-んァ-ヶー]")
HAN_RE = re.compile(r"[\u4e00-\u9fff]")
CHINESE_HINT_RE = re.compile(r"[汉把在是的了着将与并或请让给这那吗呢吧啦：；]")


def strip_number_prefix(line: str) -> str:
    return NUMBER_PREFIX_RE.sub("", line).strip()


def classify_line(line: str) -> str:
    if JAPANESE_SCRIPT_RE.search(line):
        return "japanese"
    if HAN_RE.search(line):
        return "chinese"
    if CHINESE_HINT_RE.search(line):
        return "chinese"
    if re.search(r"[A-Za-z]", line):
        return "unknown"
    return "unknown"


def likely_continuation(previous: str, current: str, language: str) -> bool:
    if not previous:
        return False
    previous_tail = previous[-1]
    if previous_tail not in "。！？!?」』）】":
        return True
    if language == "japanese" and len(previous) <= 24:
        return True
    if language == "chinese" and len(previous) <= 18:
        return True
    return False


def parse_example_block(
    block: str | None,
    *,
    row_number: int | None = None,
    max_line_length: int = 180,
) -> tuple[list[dict[str, Any]], list[WarningItem]]:
    if not block:
        return [], []

    raw_lines = [line.strip() for line in block.split("\n")]
    normalized_lines: list[tuple[int, str]] = []
    warnings: list[WarningItem] = []

    for line_number, raw_line in enumerate(raw_lines, start=1):
        line = raw_line.strip()
        if not line:
            continue
        if STANDALONE_INDEX_RE.fullmatch(line):
            warnings.append(
                WarningItem(
                    "info",
                    "overview_block",
                    "检测到单独编号行，已忽略。",
                    row_number=row_number,
                    line_number=line_number,
                )
            )
            continue
        normalized = strip_number_prefix(line)
        if len(normalized) > max_line_length:
            warnings.append(
                WarningItem(
                    "warning",
                    "overview_block",
                    f"单行长度超过 {max_line_length} 字，可能混入了解释或出现了错误换行。",
                    row_number=row_number,
                    line_number=line_number,
                )
            )
        normalized_lines.append((line_number, normalized))

    if not normalized_lines:
        return [], []

    examples: list[dict[str, Any]] = []
    pending_japanese: str | None = None
    pending_japanese_line: int | None = None
    pending_chinese: str | None = None
    pending_chinese_line: int | None = None

    def flush_example() -> None:
        nonlocal pending_japanese, pending_japanese_line, pending_chinese, pending_chinese_line
        if pending_japanese is None and pending_chinese is None:
            return
        examples.append(
            {
                "example_id": None,
                "example_order": len(examples) + 1,
                "japanese": pending_japanese,
                "chinese": pending_chinese,
                "proofreading_status": None,
                "note": None,
                "source": "overview_sheet",
                "source_row_number": row_number,
                "source_line_number": pending_japanese_line or pending_chinese_line,
            }
        )
        pending_japanese = None
        pending_japanese_line = None
        pending_chinese = None
        pending_chinese_line = None

    for line_number, line in normalized_lines:
        line_type = classify_line(line)

        if pending_japanese is None:
            if line_type == "chinese":
                warnings.append(
                    WarningItem(
                        "warning",
                        "overview_block",
                        "这里看起来是中文行，但当前位置应当开始一条例句的日文原文，可能从这一行开始配对错位。",
                        row_number=row_number,
                        line_number=line_number,
                    )
                )
            pending_japanese = line
            pending_japanese_line = line_number
            continue

        if pending_chinese is None:
            if line_type == "japanese" and likely_continuation(pending_japanese, line, "japanese"):
                pending_japanese = f"{pending_japanese}{line}"
                continue
            if line_type == "japanese":
                warnings.append(
                    WarningItem(
                        "warning",
                        "overview_block",
                        "这里仍然像日文行，但前一条例句还没有稳定读到中文翻译，可能从这一行开始配对错位。",
                        row_number=row_number,
                        line_number=line_number,
                    )
                )
            pending_chinese = line
            pending_chinese_line = line_number
            continue

        if line_type == "chinese" and likely_continuation(pending_chinese, line, "chinese"):
            pending_chinese = f"{pending_chinese}{line}"
            continue

        if line_type == "chinese":
            warnings.append(
                WarningItem(
                    "warning",
                    "overview_block",
                    "这里像是新的一行中文，但上一条例句已经结束，可能从这一行开始多拆或少拆了一行。",
                    row_number=row_number,
                    line_number=line_number,
                )
            )

        flush_example()
        pending_japanese = line
        pending_japanese_line = line_number

    if pending_japanese is not None or pending_chinese is not None:
        flush_example()

    for example in examples:
        if not example["chinese"]:
            warnings.append(
                WarningItem(
                    "warning",
                    "overview_block",
                    f"从第 {example['source_line_number']} 行开始无法稳定配对中文翻译。",
                    row_number=row_number,
                    line_number=example["source_line_number"],
                )
            )

    return examples, warnings


def normalize_overview_row(row: tuple[Any, ...], header_index: dict[str, int], row_number: int) -> dict[str, Any]:
    return {
        "row_number": row_number,
        "entry_id": clean(cell(row, header_index, "条目ID")),
        "headword": clean(cell(row, header_index, "大条目")),
        "middle_index": clean(cell(row, header_index, "中条目序号")),
        "middle_entry": clean(cell(row, header_index, "中条目")),
        "sub_index": clean(cell(row, header_index, "小条目序号")),
        "sub_entry": clean(cell(row, header_index, "小条目")),
        "meaning": clean(cell(row, header_index, "条目释义")),
        "function": clean(cell(row, header_index, "条目句型功能")),
        "connection": clean(cell(row, header_index, "条目接续")),
        "example_block_raw": clean(cell(row, header_index, "例句原文块")),
        "explanation": clean(cell(row, header_index, "条目解释")),
        "references": clean(cell(row, header_index, "参考项目")),
        "pdf_page": clean(cell(row, header_index, "PDF页码")),
        "proofreading_status": clean(cell(row, header_index, "校对状态")),
        "note": clean(cell(row, header_index, "备注")),
    }


def convert_workbook(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    missing = [name for name in (OVERVIEW_SHEET,) if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"工作簿缺少必要工作表: {', '.join(missing)}")

    overview_ws = wb[OVERVIEW_SHEET]
    overview_rows = list(overview_ws.iter_rows(values_only=True))
    overview_header = build_header_index(overview_rows[0])
    global_warnings: list[WarningItem] = []

    entries: list[dict[str, Any]] = []
    seen_entry_ids: set[str] = set()

    for row_number, row in enumerate(overview_rows[1:], start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        entry = normalize_overview_row(row, overview_header, row_number)
        entry_id = entry["entry_id"]
        if not entry_id:
            global_warnings.append(WarningItem("error", "overview_sheet", "存在缺少条目ID的主表行，已忽略。", row_number=row_number))
            continue
        if entry_id in seen_entry_ids:
            global_warnings.append(WarningItem("warning", entry_id, "主表中条目ID重复。", row_number=row_number))
        seen_entry_ids.add(entry_id)

        block_examples, block_parse_warnings = parse_example_block(entry["example_block_raw"], row_number=row_number)
        entry_warnings: list[WarningItem] = []
        for item in block_parse_warnings:
            entry_warnings.append(item if item.scope != "overview_block" else WarningItem(item.level, entry_id, item.message, item.row_number, item.line_number))

        examples = block_examples
        example_source = "overview_sheet"
        if not entry["example_block_raw"]:
            entry_warnings.append(WarningItem("info", entry_id, "当前条目没有例句。", row_number=row_number))

        entries.append(
            {
                **entry,
                "example_source": example_source,
                "examples": examples,
                "validation": {
                    "warning_count": len(entry_warnings),
                    "warnings": [item.to_dict() for item in entry_warnings],
                },
            }
        )

    usage_notes: list[dict[str, str | None]] = []
    if USAGE_SHEET in wb.sheetnames:
        usage_ws = wb[USAGE_SHEET]
        rows = list(usage_ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if not any(value is not None and str(value).strip() for value in row):
                continue
            usage_notes.append({"item": clean(row[0]), "description": clean(row[1])})

    return {
        "source_file": str(path),
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "schema_version": "excel-v1",
        "sheet_summary": {
            "overview_sheet": OVERVIEW_SHEET,
            "overview_entry_count": len(entries),
            "example_row_count": sum(len(entry["examples"]) for entry in entries),
            "example_source_mode": "overview_only",
        },
        "usage_notes": usage_notes,
        "entries": entries,
        "warnings": [item.to_dict() for item in global_warnings],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert the grammar Excel workbook into stable JSON.")
    parser.add_argument("input", type=Path, help="Path to the source .xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output JSON path. Defaults to input path with .json suffix in the same directory.",
    )
    args = parser.parse_args()

    input_path = args.input.expanduser().resolve()
    output_path = (args.output.expanduser().resolve() if args.output else input_path.with_suffix(".json"))

    data = convert_workbook(input_path)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
